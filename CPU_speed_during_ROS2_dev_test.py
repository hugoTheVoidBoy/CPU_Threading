import time
import pandas 
import psutil     #library to get CPU usage
import threading  #to run program in background while I'm working on ROS
from openpyxl import load_workbook

def get_cpu_speed():
    start = time.time()

    result = 0
    for i in range (1,1000000):
        result += i / (i+1)   #simple 1 million loops calculation as an operation to
                              #relatively time the CPU
    end = time.time()
    return end - start

def input_to_excel():
    count = 0
    activity = input("What are you gonna do today? \n")
    sheet_name = activity+ " " + time.strftime("%Hh%M'%Ss on %d-%m-%Y")
    
    while True:
        count += 1                                  #increment loop count
        cpu_speed = round(get_cpu_speed(),3)                     #get CPU speed
        cpu_usage = psutil.cpu_percent(interval=1)  #get CPU usage percent

        #start logging data
        data = {'Iteration': [count],
                'CPU speed (s)': [cpu_speed],
                'Usage (%)': [cpu_usage]
                }
        
        data_frame = pandas.DataFrame(data)
        
        #input logged data to Excel
        # try to open the existing workbook
        try:
            
            wb = load_workbook('CPU_performance_during_activities.xlsx')

            # Check if the sheet exists
            if sheet_name not in wb.sheetnames:
                # Create a new sheet if not exist
                sheet = wb.create_sheet(sheet_name)
                # Write header
                for col_num, column_title in enumerate(data_frame.columns, 1):
                    sheet.cell(row=1, column=col_num, value=column_title)

            else:
                # If the sheet exists, get the existing sheet
                sheet = wb[sheet_name]

            # Find the next empty row
            next_row = sheet.max_row + 1

            # Write the data to the that row
            for col_num, value in enumerate(data_frame.iloc[0], 1):
                sheet.cell(row=next_row, column=col_num, value=value)

            #save the workbook after appending the data
            wb.save('CPU_performance_during_activities.xlsx')
        #print error if can't open the Excel sheet
        except Exception as e:
            print(f"Error: {e}")
        
        print("Iteration ", count,". CPU speed: ", cpu_speed,". Taking over ", cpu_usage,"% of the CPU")
        if (cpu_usage > 50):
            print("OverDosing CPU. Please do some less work to prevent Overheating!")
        time.sleep(10)   # Do this every 10 secs 

def background_log():
    logging = threading.Thread(target = input_to_excel, daemon = True) #Run this in background
    logging.start()


if __name__ == "__main__":
    
    background_log()  # Start logging in the background
    print("CPU performance logging started in the background. Press Ctrl+C to stop.")
    
    try:
        while True:
            time.sleep(1)  # Main thread always work
    except KeyboardInterrupt:
        print("Logging stopped by user.")



